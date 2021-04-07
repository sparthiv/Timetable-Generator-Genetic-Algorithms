import mysql.connector as sql
import numpy as np
import random
import math
import os
import time
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet import dimensions


def cls():                          # this is to refresh screen
    if os.name == 'posix':          # if os is macos or linux
        _ = os.system('clear')
    else:
        _ = os.system('cls')        # for windows


np.set_printoptions(suppress=True)

"""The block below:
Allowing the user to enter their sql password
Connects to and fetches data from the sql database
Organises data into python data structures as required"""

sqlpass = input("Please enter password to access the database ►")
db = sql.connect(host="localhost", user="root", password=sqlpass,
                 database="ttgen", auth_plugin='mysql_native_password')
cursor = db.cursor()
str = "select count(teacher_name) from teacher_info"
cursor.execute(str)
no_of_teachers = (cursor.fetchall()[0][0])
str = "select teacher_name from teacher_info"
cursor.execute(str)
teacher_list = []
for i in range(no_of_teachers):
    teacher_list.append(cursor.fetchone()[0])

str = "select count(teachers_teaching) from class_info"
cursor.execute(str)
classes = cursor.fetchall()[0][0]
str = "select teachers_teaching from class_info"
cursor.execute(str)
teachers_for_class = []
for i in range(classes):
    teachers_for_class.append(cursor.fetchone()[0].split(","))

    for j in range(len(teachers_for_class[i])):
        teachers_for_class[i][j] = teacher_list.index(
            teachers_for_class[i][j]) + 1
str = "select class from class_info"
class_list = []
cursor.execute(str)
for i in range(classes):
    class_list.append(cursor.fetchone()[0])

popsize = 100
days = 5
periods = 5
fitness_scores = np.empty(popsize)
dimension = (classes, days, periods)
mutation = 0.005
population = np.empty([popsize, classes, days, periods])
initialpb = 0


def progressbar(l):
    """Creates a pseudo progress bar to give the user the extent of
        completetion of the algorithm. It is not a true progress
        bar as at values closer to 100% it tends to fluctuate. The
        parameter for the bar is the ratio of average fitness values
        of members of the population to the final value"""
    global initialpb
    if initialpb == 0:
        q = l
        print("|", end="")
        q = int(int(q) / 2)
        for qq in range(q):
            print("█", end='')
            time.sleep(0.1)
        for qq in range(51 - q):
            print("_", end='')
        print("|", l, "%")
        initialpb = 1
    else:
        q = l
        print("|", end="")
        q = int(int(q) / 2)
        for qq in range(q):
            print("█", end='')
        for qq in range(50 - q):
            print("_", end='')
        print("|", l, "%")


def fitness():
    """The fitness values are at the centre of the genetic algorithm. Fitness
    values are assigned to each member of the population based on which
    selection and reproduction occur. Higher the fitness value of a member of
    the population, the closer it is to the actual solution. Here the fitness
    values are calculated using a mathematical function which assigns a lower
    fitness value for greater number of collisions (same teacher having to be
    in two classes at the same time). The function ensures that even for
    small differences in the number of collisions the fitness values are far
    apart. """
.
    for i in range(popsize):
        no_of_collisions = 0
        for j in range(classes):
            temp = np.zeros(classes)
            for k in range(classes):
                temp[k] = np.count_nonzero(
                    (population[i][j] - population[i][k]) == 0)

            no_of_collisions += np.sum(temp)

        no_of_collisions -= classes * days * periods

        if no_of_collisions != 0:
            fitness_scores[i] = (no_of_collisions * math.log(0.1,
                                                             no_of_collisions)) / 1.5 + 200
        else:
            fitness_scores[i] = 200
        avg_fitness = (np.sum(fitness_scores) / popsize)
        sum_higher_fitness = 0
        count_higher = 0
        for i in range(popsize):
            if fitness_scores[i] < (avg_fitness + (
                    max(fitness_scores) + min(fitness_scores)) / 1.5) + 5:
                fitness_scores[i] = 0
            else:
                sum_higher_fitness = sum_higher_fitness + fitness_scores[i]
                count_higher = count_higher + 1
    if count_higher == 0:
        count_higher = 1
    progress = ((sum_higher_fitness / count_higher) / 2)
    progress = int(progress * 100)
    fprogress = progress / 100
    cls()
    print("PROGRESS: ")
    progressbar(fprogress)


def selection():
    a = np.array((random.choices(population, weights=fitness_scores, k=1)))
    return a[0]


def reproduce():
    """Creates the next generation which is an improvement over the
        previous generation. The parents for the next generation are
        chosen using the selection function and each member of the
        next generation is a combination of the parents. There is also
        a mutation rate which can be modified according to which values
        in the next are randomly changed so as to reach a global maximum
        for the fitness values rather than a local maxima"""
    for i in range(popsize):
        mate1 = selection()
        mate2 = selection()

        temp = np.zeros(dimension)
        for j in range(classes):
            for k in range(days):
                for l in range(periods):
                    population[i][j][k][l] = random.choice(
                        [mate1[j][k][l], mate2[j][k][l]])
                    if random.random() < mutation:
                        population[i][j][k][l] = random.choice(
                            teachers_for_class[j])


filepath = 'Timetable_Generator.xlsx'
week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']


def printtofile(arr):
    """Printing to an excel file a given number of solutions in different
    worksheets while formatting them in a readable format such as adjusting
    column size, collapsing cells, styling etc. """

    final = np.empty((classes, days, periods + 1), dtype=np.dtype(('U', 50)))
    for i in range(classes):
        for j in range(days):
            for jj in range(5):
                final[i][jj][0] = week[jj]

            for k in range(periods):
                final[i][j][k + 1] = teacher_list[int(arr[i][j][k]) - 1]
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(periods + 1):
        ws.column_dimensions[chr(65 + i)].width = 20
    for i in range(classes):
        ws.merge_cells(start_row=ws.max_row + 2, start_column=1,
                       end_row=ws.max_row + 2, end_column=periods + 1)
        cell = ws.cell(row=ws.max_row, column=1)
        cell.value = class_list[i]
        cell.style = "Title"
        cell.alignment = Alignment(vertical='center', horizontal='center')
        df = pd.DataFrame(final[i])
        for j in dataframe_to_rows(df, index=False, header=False):
            ws.append(j)
        ws.row_dimensions.group(ws.max_row - days + 1, ws.max_row, hidden=True)

    wb.save(filepath)


def initiate():
    """Creates an initial population randomly giving each class only
        the teachers that teach that class. The format of the timetables
        are Timetable>>Classes>>Days>>Periods which are stored in the
        population array"""
    for i in range(popsize):
        for j in range(classes):
            population[i][j] = np.random.choice(teachers_for_class[j],
                                                size=(days, periods),
                                                replace=True)


initiate()
start_time = time.time()
gen = 0

while True:
    """This loop runs till a solution/s is/are obtained"""
    print("CURRENT GENERATION: ", gen)

    gen = gen + 1
    fitness()
    reproduce()

    if 200 in fitness_scores:
        for i in range(popsize):
            if fitness_scores[i] == 200:
                cls()
                print("PROGRESS: 100% DONE")

                # print(population[i])
                printtofile(population[i])
                print("NUMBER OF GENERATIONS: ", gen)
                timeinsec = int(time.time() - start_time)
                if timeinsec > 59:
                    print("Elapsed time:", timeinsec // 60, "mins ",
                          timeinsec % 60, "s")
                else:
                    print(timeinsec, "s")
                os.system(filepath)

        break
